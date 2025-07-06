"""This script scrapes the interviews from Detective Conan World's Interviews page.

使用：
    python scripts/interviews_from_detectiveconanworld.py

脚本会自动：
- 爬取所有访谈内容
- 将正常访谈保存到主目录
- 将异常访谈（内容小于512字节）和汇总表保存到 abnormal_interviews 子目录
- 在Excel汇总表中用颜色标记异常情况：
  - 红色：Raw和Translation都缺失
  - 蓝色：内容异常

输出目录结构：
    data/interviews/detectiveconanworld
    ├── abnormal_interviews
    │   ├── interviews_summary.xlsx
    │   └── [异常访谈文件].txt
    └── [正常访谈文件].txt
"""

from bs4 import BeautifulSoup
import os
from datetime import datetime
import pandas as pd
import re
import openpyxl
import requests

def extract_interview(soup, section_id, index, year):
    interview = {
        "Number": index,
        "Name": "",
        "Date": "",
        "Published in": "",
        "Content": {
            "Raw": "",
            "Translation": ""
        },
        "URL": "",
        "Year": year
    }
    
    section = soup.find("span", {"id": section_id})
    if not section:
        return None
    
    interview["Name"] = section.text.strip()
    interview["URL"] = f"https://www.detectiveconanworld.com/wiki/Interviews#{section_id}"
    
    current = section.find_next()
    while current and current.name != 'div' and not (current.name == 'h3' and 'mw-headline' in current.get('class', [])):
        if current.name == 'p':
            text = current.get_text().strip()
            if 'Date:' in text:
                interview['Date'] = text.replace('Date:', '').strip()
            elif 'Published in:' in text:
                interview['Published in'] = text.replace('Published in:', '').strip()
        current = current.find_next()
    
    spoiler_div = section.find_next('div', {'id': lambda x: x and x.startswith('spoilerbordertoggledisplay')})
    if spoiler_div:
        content_div = spoiler_div.find_next('div', style=lambda x: x and 'padding: 5px; border: 1px dotted #99AACC' in x)
        if content_div:
            raw_div = content_div.find('div', {'class': 'mw-collapsible'})
            if raw_div:
                raw_content = raw_div.find('div', {'class': 'mw-collapsible-content'})
                if raw_content:
                    raw_text = [p.get_text().strip() for p in raw_content.find_all('p') if p.get_text().strip()]
                    interview['Content']['Raw'] = '\n'.join(raw_text)
            
            translation_text = []
            for p in content_div.find_all('p', recursive=False):
                text = p.get_text().strip()
                if text and not text.startswith('Raw:') and not text.startswith('Translated by:'):
                    translation_text.append(text)
            interview['Content']['Translation'] = '\n'.join(translation_text)
    
    return interview

def clean_filename(filename):
    invalid_chars = '<>:"/\\|?*'
    filename = filename.replace('/', '_').replace('\\', '_')
    for char in invalid_chars:
        if char not in ['/', '\\']:
            filename = filename.replace(char, '_')
    while '__' in filename:
        filename = filename.replace('__', '_')
    return filename.strip('_')

def extract_year(date_str):
    try:
        date_formats = [
            "%B %d, %Y",
            "%Y-%m-%d",
            "%Y"
        ]
        
        for fmt in date_formats:
            try:
                date = datetime.strptime(date_str, fmt)
                return str(date.year)
            except ValueError:
                continue
        
        year_match = re.search(r'\d{4}', date_str)
        if year_match:
            return year_match.group(0)
        
        return "Unknown"
    except:
        return "Unknown"

BASE_URL = "https://www.detectiveconanworld.com/wiki/Interviews"
DATE_FORMATS = ["%B %d, %Y", "%Y-%m-%d", "%Y"]

def build_interview_content(info):
    content = ""
    if info["Content"].get("Raw"):
        content += "Raw:\n" + str(info['Content']['Raw']) + "\n\n"
    if info["Content"].get("Translation"):
        content += "Translation:\n" + str(info['Content']['Translation'])
    return content

def save_interview(info, output_dir):
    has_raw = bool(info["Content"].get("Raw"))
    has_translation = bool(info["Content"].get("Translation"))
    
    if not has_raw and not has_translation:
        return False, has_raw, has_translation, False
    
    content = ""
    if has_raw:
        content += "Raw:\n" + str(info['Content']['Raw']) + "\n\n"
    if has_translation:
        content += "Translation:\n" + str(info['Content']['Translation'])
    content_size = len(content.encode('utf-8'))
    
    is_small_content = content_size < 512 and (has_raw or has_translation)
    
    os.makedirs(output_dir, exist_ok=True)
    abnormal_dir = os.path.join(output_dir, "abnormal_interviews")
    os.makedirs(abnormal_dir, exist_ok=True)
    
    year = info["Year"]
    filename = f"{info['Number']:03d}_{year}_{info['Name']}.txt"
    filename = clean_filename(filename)
    filepath = os.path.join(abnormal_dir if is_small_content else output_dir, filename)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(f"Name: {info['Name']}\n")
        f.write(f"Date: {info['Date']}\n")
        if info['Published in']:
            f.write(f"Published in: {info['Published in']}\n")
        f.write("\nContent:\n")
        f.write(content)
    
    return True, has_raw, has_translation, is_small_content

def main():
    output_dir = "data/interviews/detectiveconanworld"
    
    if os.path.exists(output_dir):
        print(f"Cleaning existing target directory: {output_dir}")
        import shutil
        shutil.rmtree(output_dir)
    
    os.makedirs(output_dir, exist_ok=True)
    abnormal_dir = os.path.join(output_dir, "abnormal_interviews")
    os.makedirs(abnormal_dir, exist_ok=True)
    
    url = "https://www.detectiveconanworld.com/wiki/Interviews"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    
    all_interviews_info = []
    
    print("Starting to scrape interviews...")
    try:
        print(f"Requesting webpage: {url}")
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        print("Parsing HTML content...")
        soup = BeautifulSoup(response.text, 'html.parser')
        
        print("Finding table of contents...")
        toc = soup.find("div", {"id": "toc"})
        if not toc:
            print("Error: Table of contents not found")
            return
        
        current_year = "Unknown"
        section_count = 0
        all_sections = []
        
        for item in toc.find_all("li"):
            if "toclevel-1" in item.get("class", []):
                year_text = item.find("span", {"class": "toctext"}).text.strip()
                year_match = re.search(r'\d{4}', year_text)
                current_year = year_match.group(0) if year_match else "Unknown"
            elif "toclevel-2" in item.get("class", []):
                section_count += 1
                section_id = item.find("a")["href"].lstrip("#")
                all_sections.append((section_id, section_count, current_year))
        
        print(f"Found {len(all_sections)} level-2 headings")
        
        total = len(all_sections)
        
        for i, (section_id, number, year) in enumerate(all_sections, 1):
            print(f"\nProcessing interview {i}/{total}: {section_id}")
            
            interview_info = extract_interview(soup, section_id, number, year)
            
            if interview_info:
                _, has_raw, has_translation, is_small_content = save_interview(interview_info, output_dir)
                
                interview_summary = {
                    "Number": interview_info["Number"],
                    "Year": interview_info["Year"],
                    "Title": interview_info["Name"],
                    "Date": interview_info["Date"],
                    "Published in": interview_info["Published in"],
                    "Raw exists": has_raw,
                    "Translation exists": has_translation,
                    "Abnormal content": is_small_content,
                    "Interview URL": interview_info["URL"]
                }
                all_interviews_info.append(interview_summary)
                print(f"Processed: {interview_info['Name']}")
            else:
                print(f"Warning: Failed to extract interview info: {section_id}")
        
        df = pd.DataFrame(all_interviews_info)
        excel_path = os.path.join(abnormal_dir, "interviews_summary.xlsx")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            for idx, _ in enumerate(df.index, start=2):
                has_raw = df.iloc[idx-2]['Raw exists']
                has_translation = df.iloc[idx-2]['Translation exists']
                is_small_content = df.iloc[idx-2]['Abnormal content']
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=idx, column=col)
                    
                    if not has_raw and not has_translation:
                        cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    elif is_small_content:
                        cell.fill = openpyxl.styles.PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
                    else:
                        cell.fill = openpyxl.styles.PatternFill()
        
        print(f"\nGenerated interview summary table: {excel_path}")
        print(f"Processed {len(all_interviews_info)} interviews in total")
            
    except requests.exceptions.RequestException as e:
        print(f"Request failed: {e}")
        return
    except Exception as e:
        print(f"An error occurred: {e}")
        return

if __name__ == "__main__":
    main()